#!/usr/bin/env python3
"""人工确认表 xlsx ↔ JSON 双向转换工具（KV 表/多列表格 两种 sheet）

用法:
  python3 xlsx_json_convert.py parse  <人工确认表.xlsx>           → 输出 JSON 到 stdout
  python3 xlsx_json_convert.py write  <原始人工确认表.xlsx> <out.xlsx> \
       --json-json '{"sheets": [{"name": "...", "type": "kv|table", ...}], ...}'
  或通过 stdin 传 JSON:
       cat data.json | python3 xlsx_json_convert.py write <src.xlsx> <out.xlsx>

KV 表格式 (项目概况/人员信息 两列结构):
  JSON: {"type":"kv", "rows":[{"field":"项目名称","value":"XXX"}, ...]}
  XLSX: 表头行 "字段 | 值/输入", 后续每行对应一个字段

表格型格式 (红线坐标/勘探单元/标准孔与剖线/遗迹信息/图片清单):
  JSON: {"type":"table", "headers":["A","B","C"], "rows":[["r1c1","r1c2","r1c3"], [...]]}
  XLSX: 第1行表头，后续行数据；写回时原表头不变，空行留空不写入 None
"""

import sys
import os
import json
import argparse
from typing import Any, Dict, List, Optional

KV_SHEETS = ("项目概况", "人员信息")
TABLE_SHEETS = ("红线坐标", "勘探单元", "标准孔与剖线", "遗迹信息", "图片清单")

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print("ERROR: missing openpyxl. Please: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


def _cell_value(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def parse_xlsx_to_json(xlsx_path: str) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    sheets: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        name = ws.title
        if name in KV_SHEETS:
            rows_data: List[Dict[str, str]] = []
            header_row = True
            for row in ws.iter_rows(values_only=True):
                if not row or all(c is None for c in row):
                    continue
                cells = [_cell_value(c) for c in row]
                if header_row:
                    header_row = False
                    continue
                field = cells[0] if len(cells) > 0 else ""
                value = cells[1] if len(cells) > 1 else ""
                rows_data.append({"field": field, "value": value})
            sheets.append({"name": name, "type": "kv", "rows": rows_data})
        else:
            headers: List[str] = []
            rows: List[List[str]] = []
            header_skipped = False
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                cells_raw = list(row)
                if all(c is None for c in cells_raw):
                    if not header_skipped:
                        continue
                    rows.append([""] * len(headers or cells_raw))
                    continue
                cells = [_cell_value(c) for c in cells_raw]
                if not header_skipped:
                    headers = cells
                    header_skipped = True
                    continue
                while len(cells) < len(headers):
                    cells.append("")
                rows.append(cells[: len(headers)])
            sheets.append({"name": name, "type": "table", "headers": headers, "rows": rows})
    return {"sheets": sheets}


def _write_kv_sheet(ws, rows: List[Dict[str, str]]):
    ws.cell(row=1, column=1, value="字段")
    ws.cell(row=1, column=2, value="值/输入")
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=(row.get("field") or ""))
        ws.cell(row=i, column=2, value=(row.get("value") or ""))


def _write_table_sheet(ws, headers: List[str], rows: List[List[str]]):
    col_count = len(headers)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h or "")
    for r, row in enumerate(rows, start=2):
        vals = list(row) if row is not None else []
        while len(vals) < col_count:
            vals.append("")
        for c in range(col_count):
            ws.cell(row=r, column=c + 1, value=vals[c] or "")


def write_json_to_xlsx(src_xlsx: Optional[str], out_xlsx: str, payload: Dict[str, Any]):
    wb = Workbook()
    first_sheet = wb.active
    sheets_conf: List[Dict[str, Any]] = payload.get("sheets") or []
    for idx, sheet in enumerate(sheets_conf):
        name = sheet.get("name") or f"Sheet{idx+1}"
        stype = sheet.get("type") or "table"
        if idx == 0:
            ws = first_sheet
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        if stype == "kv":
            _write_kv_sheet(ws, sheet.get("rows") or [])
        else:
            _write_table_sheet(
                ws,
                sheet.get("headers") or [],
                sheet.get("rows") or [],
            )
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)) or ".", exist_ok=True)
    wb.save(out_xlsx)


def cmd_parse(args):
    path = args.xlsx
    if not os.path.isfile(path):
        print(f"ERROR: file not found: {path}", file=sys.stderr)
        sys.exit(3)
    data = parse_xlsx_to_json(path)
    print(json.dumps(data, ensure_ascii=False, indent=2))


def cmd_write(args):
    payload: Optional[Dict[str, Any]] = None
    if args.json_json:
        payload = json.loads(args.json_json)
    else:
        payload = json.load(sys.stdin)
    if not isinstance(payload, dict) or "sheets" not in payload:
        print("ERROR: invalid JSON payload, expected {\"sheets\": [...]}", file=sys.stderr)
        sys.exit(4)
    write_json_to_xlsx(args.src, args.out, payload)
    print(json.dumps({"ok": True, "out": os.path.abspath(args.out)}, ensure_ascii=False))


def main():
    p = argparse.ArgumentParser(description="人工确认表 xlsx <-> JSON 转换")
    sub = p.add_subparsers(dest="cmd", required=True)

    pp = sub.add_parser("parse", help="解析人工确认表.xlsx -> JSON")
    pp.add_argument("xlsx", help="人工确认表 xlsx 路径")

    pw = sub.add_parser("write", help="用 JSON 覆盖原始 xlsx 内容 -> 输出新 xlsx")
    pw.add_argument("src", help="原始人工确认表路径（可填任意路径，结构以 JSON 为准保留占位即可）")
    pw.add_argument("out", help="输出 xlsx 路径")
    pw.add_argument("--json-json", default=None, help="内联 JSON 字符串（默认从 stdin 读取）")

    args = p.parse_args()
    if args.cmd == "parse":
        cmd_parse(args)
    elif args.cmd == "write":
        cmd_write(args)


if __name__ == "__main__":
    main()
